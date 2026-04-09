"""XLSX exporter for locations and characters."""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from aggregator import merge_variants


HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")


def export_xlsx(
    groups_ordered: dict | None = None,
    groups_raw: dict | None = None,
    characters: list[dict] | None = None,
) -> bytes:
    """Generate XLSX with location and/or character sheets. Returns bytes."""
    wb = Workbook()

    if groups_ordered and groups_raw:
        _write_location_sheet(wb, groups_ordered, groups_raw)

    if characters:
        _write_character_sheet(wb, characters)

    # Remove default empty sheet if we created others
    if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    elif not groups_ordered and not characters:
        # Nothing to write
        return b""

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_location_sheet(wb, groups_ordered, groups_raw):
    ws = wb.active
    ws.title = "场景库"

    headers = ["场景名(CN)", "场景名(EN)", "内景/外景", "时段", "出现集数", "场景描述", "氛围", "主要事件"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP

    row = 2
    for area_label, keys in groups_ordered.items():
        for cn, en in keys:
            scenes = groups_raw.get((cn, en), [])
            if not scenes:
                continue
            variants = merge_variants(scenes)
            for vk, v in variants.items():
                eps = sorted(v["episodes"])
                eps_str = ", ".join(f"E{e}" for e in eps)
                events_str = "\n".join(
                    f"E{e}：{ev}" for e, ev in sorted(set(v["events"]))
                )
                values = [
                    cn, en, v["int_ext"], v["time"], eps_str,
                    " ".join(v["scene_descs"]),
                    " ".join(v["atmospheres"]),
                    events_str,
                ]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.alignment = WRAP
                row += 1

    # Auto-width (approximate)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20


def _write_character_sheet(wb, characters):
    if "场景库" in wb.sheetnames:
        ws = wb.create_sheet("角色库")
    else:
        ws = wb.active
        ws.title = "角色库"

    headers = ["角色名", "中文名", "性别", "年龄", "身份", "外貌描述", "性格", "出场集数", "主要关系", "角色层级"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP

    for i, c in enumerate(characters, 2):
        eps = c.get("episodes", [])
        eps_str = ", ".join(f"E{e}" for e in sorted(eps)) if eps else ""
        rels = c.get("key_relationships", [])
        rels_str = "\n".join(
            f"{r.get('character', '?')}：{r.get('relationship', '')}" for r in rels
        )
        values = [
            c.get("character_name", ""),
            c.get("character_name_cn", ""),
            c.get("gender", ""),
            c.get("age", ""),
            c.get("identity", ""),
            c.get("appearance", ""),
            c.get("personality", ""),
            eps_str,
            rels_str,
            c.get("role_tier", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = WRAP

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

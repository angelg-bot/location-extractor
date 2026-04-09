"""XLSX → plain text using openpyxl."""

from openpyxl import load_workbook


def parse_xlsx(file_path: str, sheet_name: str = None) -> str:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        sheets = [wb[sheet_name]]
    else:
        sheets = [wb[wb.sheetnames[0]]]

    lines = []
    for ws in sheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = "\t".join(cells).strip()
            if line:
                lines.append(line)

    wb.close()
    return "\n".join(lines)
